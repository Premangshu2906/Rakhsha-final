import os
import re
import json
import logging
from typing import Dict, List, Tuple, Any

logger = logging.getLogger("NHAA_DatasetLoader")

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
# Fallback if BASE_DIR isn't NHAA-Portal
if not os.path.exists(os.path.join(BASE_DIR, "fearful_words_danger_dataset_english.xlsx")):
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

ENGLISH_DATASET_PATH = os.path.join(BASE_DIR, "fearful_words_danger_dataset_english.xlsx")
HINGLISH_DATASET_PATH = os.path.join(BASE_DIR, "RAKHSHA_Hinglish_Danger_Signal_Dataset.xlsx")
CACHE_FILE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "trained_lexicon.json")

class DatasetTrainedModel:
    """
    Trained Lexicon & Danger Scoring Engine built on:
    1. English Fearful Words Dataset (578 words)
    2. RAKHSHA Hinglish Danger Signal Dataset (310 phrases)
    Total: 888 rated vocabulary entries across Scores 1 (Mild) to 5 (Critical).
    """

    SEVERITY_MAP = {
        1: "Mild",
        2: "Moderate",
        3: "Serious",
        4: "Severe",
        5: "Critical"
    }

    def __init__(self):
        self.english_lexicon: Dict[str, Dict[str, Any]] = {}
        self.hinglish_lexicon: Dict[str, Dict[str, Any]] = {}
        self.phrase_patterns: List[Tuple[re.Pattern, str, int, str]] = []
        self.is_loaded = False
        self.load_datasets()

    def load_datasets(self):
        """Loads dataset entries from Excel files or fallback cache JSON."""
        loaded_from_excel = False

        try:
            import openpyxl
            
            # Load English Dataset
            if os.path.exists(ENGLISH_DATASET_PATH):
                wb1 = openpyxl.load_workbook(ENGLISH_DATASET_PATH, data_only=True)
                ws1 = wb1['word_scores']
                for row in list(ws1.iter_rows(values_only=True))[1:]:
                    if row and len(row) >= 2 and row[0]:
                        word = str(row[0]).strip().lower()
                        score = int(row[1]) if row[1] is not None else 1
                        severity = str(row[2]) if len(row) > 2 and row[2] else self.SEVERITY_MAP.get(score, "Mild")
                        self.english_lexicon[word] = {
                            "word": word,
                            "score": score,
                            "severity": severity,
                            "lang": "English"
                        }
                logger.info(f"Loaded {len(self.english_lexicon)} English fearful words from Excel.")

            # Load Hinglish Dataset
            if os.path.exists(HINGLISH_DATASET_PATH):
                wb2 = openpyxl.load_workbook(HINGLISH_DATASET_PATH, data_only=True)
                ws2 = wb2['Hinglish Dataset']
                for row in list(ws2.iter_rows(values_only=True))[1:]:
                    if row and len(row) >= 3 and row[1]:
                        phrase = str(row[1]).strip().lower()
                        score = int(row[2]) if row[2] is not None else 1
                        lang = str(row[3]) if len(row) > 3 and row[3] else "Hinglish"
                        self.hinglish_lexicon[phrase] = {
                            "word": phrase,
                            "score": score,
                            "severity": self.SEVERITY_MAP.get(score, "Mild"),
                            "lang": lang
                        }
                logger.info(f"Loaded {len(self.hinglish_lexicon)} Hinglish danger phrases from Excel.")

            # Devanagari Hindi Threat Vocabulary Mapping for direct Hindi Speech-to-Text outputs
            devanagari_dict = {
                "जान से मार": (5, "Critical"), "मार डालूंगा": (5, "Critical"), "मार देंगे": (5, "Critical"),
                "मार दूंगा": (5, "Critical"), "हत्या": (5, "Critical"), "जान का खतरा": (5, "Critical"),
                "तेजाब": (5, "Critical"), "एसिड": (5, "Critical"), "आत्महत्या": (5, "Critical"),
                "फांसी": (5, "Critical"), "जहर": (5, "Critical"), "ज़हर": (5, "Critical"),
                "गोली": (5, "Critical"), "बंदूक": (5, "Critical"), "मारपीट": (4, "Severe"),
                "मारना": (4, "Severe"), "पीट": (4, "Severe"), "पीटा": (4, "Severe"),
                "चाकू": (4, "Severe"), "तलवार": (4, "Severe"), "खून": (4, "Severe"),
                "घायल": (4, "Severe"), "जला": (4, "Severe"), "कमरे में बंद": (4, "Severe"),
                "ताला": (4, "Severe"), "बंधक": (4, "Severe"), "धमकी": (3, "Serious"),
                "खतरा": (3, "Serious"), "डर": (3, "Serious"), "गाली": (3, "Serious"),
                "गालियां": (3, "Serious"), "बचाओ": (3, "Serious"), "मदद": (3, "Serious"),
                "परेशान": (3, "Serious"), "प्रताड़ित": (3, "Serious"), "दहेज": (3, "Serious"),
                "ब्लैकमेल": (3, "Serious"), "जाति": (3, "Serious"), "अपमान": (3, "Serious"),
                "झगड़ा": (2, "Moderate"), "छेड़छाड़": (2, "Moderate"), "शिकायत": (1, "Mild")
            }

            for dev_phrase, (score, severity) in devanagari_dict.items():
                self.hinglish_lexicon[dev_phrase] = {
                    "word": dev_phrase,
                    "score": score,
                    "severity": severity,
                    "lang": "Hindi (Devanagari)"
                }

            if self.english_lexicon or self.hinglish_lexicon:
                loaded_from_excel = True
                self._save_cache()

        except Exception as e:
            logger.warning(f"Excel dataset loading warning: {e}. Falling back to cache.")

        if not loaded_from_excel:
            self._load_cache()

        self._build_compiled_patterns()
        self.is_loaded = True

    def _save_cache(self):
        """Saves processed lexicon to JSON cache for fast loading."""
        try:
            cache_data = {
                "english": self.english_lexicon,
                "hinglish": self.hinglish_lexicon,
                "total_entries": len(self.english_lexicon) + len(self.hinglish_lexicon)
            }
            with open(CACHE_FILE_PATH, "w", encoding="utf-8") as f:
                json.dump(cache_data, f, indent=2, ensure_ascii=False)
            logger.info(f"Saved trained lexicon cache ({cache_data['total_entries']} entries) to {CACHE_FILE_PATH}")
        except Exception as e:
            logger.error(f"Error saving lexicon cache: {e}")

    def _load_cache(self):
        """Loads from JSON cache file."""
        if os.path.exists(CACHE_FILE_PATH):
            try:
                with open(CACHE_FILE_PATH, "r", encoding="utf-8") as f:
                    cache_data = json.load(f)
                self.english_lexicon = cache_data.get("english", {})
                self.hinglish_lexicon = cache_data.get("hinglish", {})
                logger.info(f"Loaded {cache_data.get('total_entries', 0)} entries from cache JSON.")
            except Exception as e:
                logger.error(f"Error reading cache JSON: {e}")

    def _build_compiled_patterns(self):
        """Compiles regex patterns sorted by phrase length (longest first) for optimal multi-word matching."""
        self.phrase_patterns = []
        
        # Combine all phrases
        combined = []
        for phrase, data in self.hinglish_lexicon.items():
            combined.append((phrase, data["score"], data["severity"], data["lang"]))
        for word, data in self.english_lexicon.items():
            combined.append((word, data["score"], data["severity"], data["lang"]))

        # Sort by length descending so longer multi-word phrases match before individual words
        combined.sort(key=lambda x: len(x[0]), reverse=True)

        for text_str, score, severity, lang in combined:
            # Escape text for regex matching with word boundaries
            escaped = re.escape(text_str)
            pattern = re.compile(r'\b' + escaped + r'\b', re.IGNORECASE)
            self.phrase_patterns.append((pattern, text_str, score, severity, lang))

    def evaluate_text(self, text: str) -> Dict[str, Any]:
        """
        Evaluates input text against both English and Hinglish datasets.
        Returns:
        - matched_indicators: List of matching entries with word, score, severity, language
        - max_danger_score: Maximum danger score detected (1-5)
        - weighted_score: Continuous calculated distress score (0.0 to 100.0)
        - hinglish_detected: True if Hinglish phrases were detected
        - english_detected: True if English danger words were detected
        """
        text_lower = text.lower()
        matched_indicators = []
        matched_words_set = set()

        max_score = 0
        total_score_sum = 0
        score_5_count = 0
        score_4_count = 0
        score_3_count = 0

        hinglish_detected = False
        english_detected = False

        for pattern, text_str, score, severity, lang in self.phrase_patterns:
            if text_str in matched_words_set:
                continue
            if pattern.search(text_lower):
                matched_words_set.add(text_str)
                matched_indicators.append({
                    "phrase": text_str,
                    "danger_score": score,
                    "severity": severity,
                    "language": lang
                })
                max_score = max(max_score, score)
                total_score_sum += score

                if score == 5:
                    score_5_count += 1
                elif score == 4:
                    score_4_count += 1
                elif score == 3:
                    score_3_count += 1

                if lang == "Hinglish":
                    hinglish_detected = True
                else:
                    english_detected = True

        # Calculate continuous distress score (0-100)
        # Score calculation formula based on dataset weights
        base_score = 0.0
        if score_5_count > 0:
            base_score = 75.0 + min(score_5_count * 8.0, 25.0)
        elif score_4_count > 0:
            base_score = 55.0 + min((score_4_count * 10) + (score_3_count * 5), 25.0)
        elif score_3_count > 0:
            base_score = 35.0 + min(score_3_count * 10, 20.0)
        elif max_score == 2:
            base_score = 25.0 + min(len(matched_indicators) * 5, 15.0)
        elif max_score == 1:
            base_score = 15.0 + min(len(matched_indicators) * 3, 10.0)

        weighted_score = round(min(base_score, 100.0), 1)

        return {
            "matched_indicators": matched_indicators,
            "matched_count": len(matched_indicators),
            "max_danger_score": max_score,
            "distress_score": weighted_score,
            "hinglish_detected": hinglish_detected,
            "english_detected": english_detected,
            "critical_matches": score_5_count + score_4_count
        }

# Global singleton model instance
dataset_model = DatasetTrainedModel()
